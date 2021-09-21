#!/usr/bin/python3
import os
import copy

import pandas as pd
import numpy as np
from openpyxl import load_workbook # Write Excel Files

def has_apostrophe(text):
    if type(text) is not float:
        return text.replace("'", "")
    return text

def get_loyalty_discounts_refunds(row):
    loyalty, discount, refund = 0, 0, 0
    
    unit_price = row['Order Lines/Unit Price']
    qty = row['Order Lines/Quantity']
    product_name = row['Order Lines/Product/Name']
    product_cat = row['Order Lines/Product/Product Category']
    
    if unit_price < 0:
        if ('Loyalty' in product_name) & ('Discount' in product_cat):
            loyalty = -1.0 * unit_price * qty
        else:
            discount = -1.0 * unit_price * qty
    elif qty < 0:
        refund = -1.0 * unit_price * qty
        
    return (loyalty, discount, refund)
    
def simple_payment_names(text):
    basic_types = ['Credit', 'Cash', 'Card']
    sub_types = ['Talabat', 'Insta', 'Carriage',
                 'Baqalaat', 'BEC', 'Jebly', 'PKG',
                 'HOMIEZ', 'Note', 'Henlo', 'Feehla', 'Dalooni' ]
    simple_name = []
    for typ in basic_types:
        if typ in text:
            simple_name.append(typ)
            break
    
    for typ in sub_types:
        if typ in text:
            simple_name.append(typ)
            break
    
    return '|'.join(simple_name)
        
def get_type_of_sale(row):

    if row['Order Lines/Product/Product Category'] == 'Grooming Service':
        return 'Grooming'

    product_types = ['Cash', 'Card', 'Credit|Note']
    if row['Payments/Journal/Journal Name'] in product_types:
        return 'Product'
    return row['Payments/Journal/Journal Name']

def get_pivot_table(df, columns, values, rows):
    agg_cols = ['Loyalty Points', 'TotalDiscount', 'Refunds', 'Total']
    agg_part = df[rows+agg_cols].groupby(rows).sum().reset_index()
    pivot_table = pd.pivot(df, index=rows, columns=columns,
                        values=values).reset_index()
    # Merge Them
    return pd.merge(pivot_table,agg_part,how='inner',on=rows)

def fill_in_missing_dates(df, month_as_date):
    start_date = month_as_date
    end_date = str(pd.date_range(start=start_date, periods=1, freq='M', ).item())[:10]
    date_series = pd.date_range(start=start_date, end=end_date, freq='D')
    date_series = pd.Series(date_series)

    date_series.name = 'Date'


    date_series = date_series.dt.strftime("%Y-%m-%d")
    output_df = df.merge(date_series, how='right',left_on='Date', right_on='Date')\
                  .fillna(0.)
    return output_df

def insert_xltable(ws, df, df_to_xlmap):
    # Create Columns if they don't exist
    for (start_row, start_col, header), cols in df_to_xlmap.items():
        subset = df[cols].values
        if header is True:
            subset = np.vstack([cols, subset])
        
        width = subset.shape[1]
        height = subset.shape[0]
            
        for row in range(start_row, start_row+height):
            for col in range(start_col, start_col+width):
                value = subset[row-start_row, col-start_col]
                try:
                    value = float(value)
                except:
                    pass
                ws.cell(row=row, column=col,
                        value=value)

def select_df_bymonth(df, yearmonth):
    # Example : yearmonth = "2021-07"
    yearmonth += '-01'

    end_date = str(pd.date_range(start=yearmonth, periods=1, freq='M', ).item())[:10]
    date_series = pd.date_range(start=yearmonth, end=end_date, freq='D')
    date_series = pd.Series(date_series)

    date_series.name = 'Date'

    # Generate the date series
    date_series = date_series.dt.strftime("%Y-%m-%d")

    # Select the Month
    month_df = df[(df['Date'] >= yearmonth) & (df['Date'] <= end_date)]

    # Merge them
    output_df = month_df.merge(date_series, how='right',left_on='Date', right_on='Date')\
                        .fillna(0.)
    return output_df

def generate_report(input_csv, dest_dir, yearmonth, prefix_filename = 'DSR_'):

    # Read the CSV
    df = pd.read_csv(input_csv)
    # Drop Null Values
    df.dropna(axis=0, how='any', subset=['Order Lines/Subtotal'], inplace=True)

    # Fill Remaining Null Values

    # Convert Numerical Types
    float_cols = ['Order Lines/Subtotal', 'Total', 'Order Lines/Unit Price']
    int_cols = ['Order Lines/Quantity']
    df[float_cols+int_cols] = df[float_cols+int_cols].applymap(has_apostrophe).astype('float64')

    # Rows with NA values in Order Lines Fields are removed
    numbers_cols = ["Order Lines/Unit Price","Order Lines/Quantity","Order Lines/Discount (%)","Order Lines/Discount Fixed","Order Lines/Subtotal"]
    df[numbers_cols] = df[numbers_cols].fillna(0.)

    df.fillna(method='ffill', inplace=True)
    df[int_cols] = df[int_cols].astype('int32')

    ## Removed Bathclub Products
    indices = df['Order Lines/Product/Name'].map(lambda x : 'Gift-Coupon' in x)
    df = df[~indices]
    indices = df['Order Lines/Product/Product Category'].map(lambda x : 'Service-Bathclub' in x)
    df = df[~indices]

    ## Loyalty Points, Discounts, Refunds
    df[['Loyalty Points','TotalDiscount', 'Refunds']] = df.apply(get_loyalty_discounts_refunds, result_type='expand',axis=1)

    ## Finding Real Total Discount
    # % Discount
    df['TotalDiscount'] += df['Order Lines/Unit Price'] * df['Order Lines/Quantity'] * (df['Order Lines/Discount (%)']/100)
    # Fixed Discount 
    df['TotalDiscount'] += df['Order Lines/Discount Fixed']

    ## Create New Columns
    df['Branch'] = df['Point of Sale Name'].dropna().map(lambda x: x.split(' ')[0])

    # Trim Dates
    df['Order Date'] = df['Order Date'].map(lambda x: x[:10])

    ## Simplify Payment Method Names
    df['Payments/Journal/Journal Name'] = df['Payments/Journal/Journal Name'].map(simple_payment_names)


    # ####### SEPARATE SALE TYPES : Grooming, Products, App Sales
    df['SaleType'] = df.apply(get_type_of_sale, axis=1)


    # Remove Redundant Columns
    df.drop(['External ID', 'Point of Sale Name', 'Status', 'Total', 'Customer/Name', 'Order Lines/Product/Name', 'Order Lines/Product/Product Category', 'Order Lines/Unit Price', 'Order Lines/Quantity',
         'Order Lines/Discount (%)', 'Order Lines/Discount Fixed'],axis=1,inplace=True)
    
    # ####### GROUPBY DataFrames 
    multi_indices = ['Order Date', 'Branch', 'Salesperson/Name', 'Payments/Journal/Journal Name', 'SaleType']
    order_totals = pd.DataFrame(df.groupby(by=multi_indices).sum()).reset_index()

    # Rename Columns
    order_totals.rename({'Salesperson/Name': 'Salesperson',
                      'Order Date': 'Date',
                      'Payments/Journal/Journal Name': 'PaymentType',
                     'Order Lines/Subtotal': 'Total'}, axis=1, inplace=True)
    
    ## Finally Separate : App Vs. Grooming Vs. Product
    app_order_indices = order_totals['SaleType'].map(lambda typ: typ not in ['Product', 'Grooming'])
    app_orders = order_totals[app_order_indices].copy()

    # Shop Orders - > Product + Grooming
    shop_orders = order_totals[~app_order_indices].copy()
    product_orders = shop_orders[shop_orders['SaleType'] == 'Product'].copy()
    groom_orders = shop_orders[shop_orders['SaleType'] == 'Grooming'].copy()

    product_orders.drop('SaleType', inplace=True, axis=1)
    groom_orders.drop('SaleType', inplace=True, axis=1)
    app_orders.drop('SaleType', inplace=True, axis=1)

    ## PIVOTING TABLES
    indices = ['Branch','Date','Salesperson']

    products_table = get_pivot_table(product_orders, ['PaymentType'], 'Total', indices)
    groom_table = get_pivot_table(groom_orders, ['PaymentType'], 'Total', indices)
    app_table = get_pivot_table(app_orders, ['PaymentType'], 'Total', indices)

    # Merge Them All
    merg1 = pd.merge(products_table, groom_table, how='outer', on=indices, suffixes=('_P', '_G'))
    whole_table = pd.merge(merg1, app_table, how='outer', on=indices, suffixes=('_S', '_App'))
    whole_table.fillna(0., inplace=True)

    # Remove SalesPerson
    whole_table = whole_table.groupby(['Branch', 'Date']).sum().reset_index()


    #### BUILD EXCEL FILE

    # Mapping between row, column indices in the Excel File to the fields in the DataFrame
    excel_sheet_map = { 
    # (Row, Col, NeedHeader?): Index,
    (4,1, False): ['Date'],
    (4,4, False): ['Cash_P', 'Card_P', 'Credit|Note_P', 'Loyalty Points_P', 'TotalDiscount_P', 'Refunds_P'], #Product Sales
    (4,12, False): ['Cash_G', 'Card_G', 'Credit|Note_G', 'Loyalty Points_G', 'TotalDiscount_G', 'Refunds_G'], # Grooming Sales
    (3,26, True): app_orders['PaymentType'].unique().tolist() + ['Total']
    }
    # Create columns if they don't exist
    for cols in excel_sheet_map.values():
        for c in cols:
            if c not in whole_table.columns:
                whole_table[c] = 0.0


    # Load Excel Template
    wb = load_workbook(filename = 'Branch_Daily_Sales_Report_Sample.xlsx')

    for branch in whole_table.Branch.unique():
        branch_orders = whole_table[whole_table['Branch'] == branch]

        # Select Month & Fill in missing Days with Zeros
        month_orders = select_df_bymonth(branch_orders, yearmonth) 


        wb_copy = copy.copy(wb)
        ws = wb_copy['JUL']
        # ws.title = 'REPORT'
        insert_xltable(ws, month_orders, excel_sheet_map)
        save_path = os.path.join(dest_dir, f"{prefix_filename}{branch}.xlsx")
        wb_copy.save(save_path)


if __name__ == '__main__':

    generate_report('pos.order2020-1.csv', 'output')
    print("\nDone!")